"""ورود گروهی محصولات فروشگاه از Excel/CSV."""

from __future__ import annotations

import csv
import io
from typing import Any

from balebot.services.slug_utils import resolve_model_slug, slug_from_text
from openpyxl import Workbook, load_workbook

from balebot.services.catalog_currency import toman_to_rial

REQUIRED_COLUMNS = ('name', 'category', 'price')
OPTIONAL_COLUMNS = ('stock', 'description', 'image_url')
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


def _normalize_header(value: str) -> str:
    return (value or '').strip().lower().replace(' ', '_')


def _parse_row_dict(row: dict[str, Any], line_no: int) -> tuple[dict[str, Any] | None, str | None]:
    name = (row.get('name') or '').strip()
    category = (row.get('category') or '').strip()
    price_raw = row.get('price')
    if not name:
        return None, f'ردیف {line_no}: نام محصول خالی است.'
    if not category:
        return None, f'ردیف {line_no}: دسته‌بندی خالی است.'
    try:
        price = int(float(str(price_raw).replace(',', '').strip()))
    except (TypeError, ValueError):
        return None, f'ردیف {line_no}: قیمت نامعتبر است.'
    if price < 0:
        return None, f'ردیف {line_no}: قیمت نمی‌تواند منفی باشد.'

    stock_val = None
    stock_raw = row.get('stock')
    if stock_raw not in (None, ''):
        try:
            stock_val = int(float(str(stock_raw).strip()))
        except (TypeError, ValueError):
            return None, f'ردیف {line_no}: موجودی نامعتبر است.'

    return {
        'name': name[:200],
        'category': category[:120],
        'price': price,
        'stock': stock_val,
        'description': (row.get('description') or '')[:5000],
        'image_url': (row.get('image_url') or '').strip()[:500],
        'slug': slug_from_text(name, fallback=f'item-{line_no}', max_length=220),
    }, None


def _rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    header_map = {_normalize_header(h): h for h in reader.fieldnames if h}
    rows: list[dict[str, Any]] = []
    for raw in reader:
        row = {_col: raw.get(header_map[_col], '') for _col in ALL_COLUMNS if _col in header_map}
        if any(str(v).strip() for v in row.values()):
            rows.append(row)
    return rows


def _rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [_normalize_header(str(h or '')) for h in header_row]
    idx = {name: i for i, name in enumerate(headers) if name}
    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        if not values or not any(v not in (None, '') for v in values):
            continue
        row: dict[str, Any] = {}
        for col in ALL_COLUMNS:
            if col in idx:
                val = values[idx[col]]
                row[col] = '' if val is None else val
        rows.append(row)
    return rows


def parse_import_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower = (filename or '').lower()
    if lower.endswith('.csv'):
        return _rows_from_csv(content)
    if lower.endswith('.xlsx') or lower.endswith('.xls'):
        return _rows_from_xlsx(content)
    raise ValueError('فقط فایل Excel (.xlsx) یا CSV پشتیبانی می‌شود.')


def preview_import_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    valid: list[dict[str, Any]] = []
    errors: list[str] = []
    for i, raw in enumerate(rows, start=2):
        parsed, err = _parse_row_dict(raw, i)
        if err:
            errors.append(err)
        elif parsed:
            valid.append(parsed)
    return valid, errors


def build_sample_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'products'
    ws.append(list(ALL_COLUMNS))
    ws.append(['مانتو کرپ', 'مانتو', 385000, 10, 'توضیح نمونه', ''])
    ws.append(['شال نخی', 'اکسسوری', 69000, '', 'نخ ابریشم', ''])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_rows(
    catalog: CatalogSettings,
    rows: list[dict[str, Any]],
) -> dict[str, int]:
    created_items = 0
    created_categories = 0
    workspace = catalog.workspace
    platform = catalog.platform

    for row in rows:
        cat_name = row['category']
        cat_slug = slug_from_text(cat_name, fallback='cat', max_length=140)
        category, cat_created = CatalogCategory.objects.get_or_create(
            workspace=workspace,
            platform=platform,
            slug=cat_slug,
            defaults={'name': cat_name, 'is_active': True},
        )
        if cat_created:
            created_categories += 1

        slug = resolve_model_slug(
            CatalogItem,
            row['name'],
            workspace=workspace,
            platform=platform,
            fallback='item',
            max_length=220,
        )

        CatalogItem.objects.create(
            workspace=workspace,
            platform=platform,
            category=category,
            title=row['name'],
            slug=slug,
            description=row.get('description') or '',
            price=toman_to_rial(row['price']),
            stock=row.get('stock'),
            item_type=CatalogItem.ItemType.PRODUCT,
            sale_mode=CatalogItem.SaleMode.BUYABLE,
            is_active=True,
        )
        created_items += 1

    return {
        'items_created': created_items,
        'categories_created': created_categories,
    }
