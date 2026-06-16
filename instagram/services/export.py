from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

PHONE_EXPORT_HEADERS = ('شماره', 'حوزه فعالیت', 'تاریخ', 'شناسه عملیات')
JOB_PHONE_EXPORT_HEADERS = ('شماره', 'حوزه فعالیت', 'تاریخ')


def build_phones_xlsx(headers: tuple[str, ...], rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'شماره\u200cها'
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True)
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = header_font

    for row in rows:
        ws.append(list(row))

    for idx, _ in enumerate(headers, start=1):
        column = get_column_letter(idx)
        max_len = max(
            len(str(ws[f'{column}{r}'].value or ''))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[column].width = min(max(max_len + 2, 12), 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
